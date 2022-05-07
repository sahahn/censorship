import json
import branca
import folium
import time
import pandas as pd
import os
import openpyxl
import base64
from collections import defaultdict


def wget(link, file_name=None):
  '''Just wrap in python to avoid downloading more times than needed'''

  # Skip if already downloaded
  if file_name is None:
    file_name = link.split('/')[-1]

  if os.path.exists(file_name):
    return

  # Call wget to download
  os.system(f'wget -O {file_name} {link}')

def dl():

    wget('https://cdn.jsdelivr.net/npm/us-atlas@3/states-10m.json')
    wget('https://docs.google.com/spreadsheets/d/1Tj5WQVBmB6SQg-zP_M8uZsQQGH09TxmBY73v23zpyr0/export?format=xlsx&gid=107383712', 'latest.xlsx')
    wget('https://raw.githubusercontent.com/jasonong/List-of-US-States/master/states.csv')
    wget('https://raw.githubusercontent.com/benkeen/miscellaneous/master/d3collision/us-state-centroids.json')

    # Make sure downloaded before proceeding?
    time.sleep(10)

def _proc_year(x):
  
  base = str(x).split('\n')[0].rstrip()

  try:
    return int(str(base).split('-')[0])
  except:
    return int(str(base).split('/')[-1])

def load_base_df():

  try:
    df = pd.read_excel('latest.xlsx', sheet_name=3)
  except:
    df = pd.read_excel('latest.xlsx', sheet_name=0)

  print(list(df))
  
  df = df.drop(columns = ['Primary Sponsor'])

  # Clean the column names
  df.columns = [x.replace(' ', '_') for x in df.columns]

  # Renaming the terribly named ones
  cols = list(df)
  df = df.rename({cols[1]: 'Bill_Num',
                  cols[2]: 'Year',
                  cols[3]: 'Status',
                  cols[4]: 'Description',
                  cols[6]: 'Enforcement/Penalties'}, axis=1)
  
  # Convert year from date
  df['Year'] = df['Year'].apply(lambda x: _proc_year(x))
  
  return df


def check_target_key_words(row, *keys):

  # Idea is to check if any of the key
  # words are in descr or explicitly targets
  for key in keys:
    if key in row['Description'] or key in row['Explicitly_Targets']:
      return 1
  
  # Not found
  return 0

def add_bill_targets(df):

  # Change to lowercase to make looking for keywords easier
  df['Description'] = [x.lower() for x in df['Description']]
  df['Explicitly_Targets'] = [x.lower() for x in df['Explicitly_Targets']]

  # Each columns / row's value reflects if any of those keywords
  # are present in either the description or the explicitly targets column
  # keywords to check are passed as args
  df['K12'] = df.apply(check_target_key_words, args=['k-12', 'k12'], axis=1)
  df['Colleges'] = df.apply(check_target_key_words, args=['college', 'colleges'], axis=1)
  df['State Contractors / Institutions'] = df.apply(check_target_key_words, args=['state agencies', 'state contractors'], axis=1)
  df['Gender / Sexuality'] = df.apply(check_target_key_words, args=['gender', 'sex', 'sexuality'], axis=1)
  df['Religion'] = df.apply(check_target_key_words, args=['religion'], axis=1)
  df['Race / Ethnicity'] = df.apply(check_target_key_words, args=['race', 'ethnicity', 'ethnic'], axis=1)

  return df

def simplify_status(status):

  status = status.lower()

  died_keys = ['died', 'vetoed', 'withdrawn', 'struck down', 'ruled out of order']
  signed_keys = ['passed', 'signed', 'became law']
  pending_keys = ['prefiled', 'pending']

  for key in died_keys:
    if key in status:
      return 'Died'
  for key in signed_keys:
    if key in status:
      return 'Signed'
  for key in pending_keys:
    if key in status:
      return 'Pending'

  # Raise error if none
  raise RuntimeError(f'No status for: {status}')

def add_status_cols(df):
  
  # Refine status col
  df['Refined Status'] = df['Status'].apply(simplify_status)

  # Add cols based on refined status
  df['Num Pending'] = df['Refined Status'] == 'Pending'
  df['Num Died'] = df['Refined Status'] == 'Died'
  df['Num Signed'] = df['Refined Status'] == 'Signed'

  # Identity col
  df['Num Bills'] = 1

  return df

# Put everything so far together
# So can call this whenever and get 
# a clean copy
def get_df():

  df = load_base_df()
  df = add_bill_targets(df)
  df = add_status_cols(df)

  return df

def remove_extra_cols(df, extra):

  # Skip if already droped
  if 'Explicitly_Targets' not in df:
    return df

  # Base columns that we don't want to plot
  to_drop = ['Explicitly_Targets', 'Enforcement/Penalties', 'Year',
             'Description','Bill_Num', 'Status', 'Refined Status']

  # If grouping by another column
  # don't drop
  if extra is not None:
    if extra in to_drop:
      to_drop.remove(extra)

  return df.drop(columns=to_drop)

def sum_df_by_state(df, extra=None):

  # Remove extra cols
  summed_df = remove_extra_cols(df, extra=extra)

  # Sum
  if extra is None:
    summed_df = summed_df.groupby('State').apply(sum)
  else:
    summed_df = summed_df.groupby([extra, 'State']).apply(sum)

  # Add abr col
  # summed_df['State Abr'] = [us_state_to_abbrev[i.strip()] for i in states]

  # Not sure why it keeps the group by columns as columns
  if 'State' in summed_df:
    summed_df = summed_df.drop('State', axis=1)
  if extra in summed_df:
    summed_df = summed_df.drop(extra, axis=1)

  return summed_df

def to_hover_full_html(row, state_name):

  if isinstance(row, pd.Series):
    table_html = row.to_frame().to_html(header=False)
  else:
    table_html = row.to_html()

  html = """
<!DOCTYPE html>
<html>
<style>
table {
  border-collapse: collapse;
  width: 100%;
}

th, td {
  padding: 8px;
  text-align: left;
  border-bottom: 1px solid #ddd;
}
</style>""" +\
"""<center><h4 style="margin-bottom:5"; width="200px">{}</h4>""".format(state_name) + """</center>
"""+table_html+"""
</html>
"""
  return html


def load_geo(ref_df):
  
  # Load base
  with open('states-10m.json', 'r') as f:
    geo = json.load(f)

  # This is where we add the content for the hover-text
  # I.e., any fields we want to reference in the hover text
  # need to be added here under properties for each state
  for s in geo['objects']['states']['geometries']:
    
    # The name of the current state
    state_name = s['properties']['name']
    
    # Get the row from the base summed all df
    # if doesn't exist, skip
    try:
      row = ref_df.loc[state_name]
      s['properties']['info'] = to_hover_full_html(row, state_name)
    except KeyError:
      s['properties']['info'] = 'No bills introduced!'

  return geo

# Style function should change based on reference dataframe
# e.g., either base or summed by state x other column
def get_style_function(summed_df, col, colorscale):

  def style_function(feature):
    '''Method used to map colors to locations.
    This function is called by the TopoJson method for each state
    and is used to determine the color.'''

    keys = list(feature["properties"].keys())
    name = feature["properties"][keys[0]]

    # Get value from summed_df w/ error handling
    try:
      row = summed_df.loc[name]
      count = row[col]
    except:
      count = None
    
    # Convert rank to color
    color = "#black" if count is None else colorscale(count)
    fill_o = 0 if count is None else .666

    return {"fillOpacity": fill_o,
            "weight": 0,
            "fillColor": color}

  return style_function

def load_links():
    
  # Need to load with openpyxl to retain hyper-links
  wb = openpyxl.load_workbook('latest.xlsx')

  # Get's all sheets
  sheets = wb.sheetnames

  if len(sheets) == 1:
    sheet = 0
  else:
    sheet = 3

  # Select the passed sheet by int index
  ws = wb[sheets[sheet]]
  
  # Populate a dictionary of links
  # where dict key corresponds to int index
  # in the table
  links = {}
  for i in range(1, ws.max_row+1):
      
      try:
          state = ws.cell(row=i, column=1).value
          link = ws.cell(row=i, column=2).hyperlink.target
          name = ws.cell(row=i, column=2).value
          
          if state in links:
            links[state].append((name, link))
          else:
            links[state] = [(name, link)]
          
      # Skip any without links
      except AttributeError:
          pass

  return links

def to_popup_full_html(state_name, links):

  bills_html_links = ''.join([f'<li><a href="{link[1]}" target="_blank">{link[0]}</a></li>' for link in links[state_name]])

  html = """
<!DOCTYPE html>
<html>
<style>
table {
  border-collapse: collapse;
  width: 100%;
}

th, td {
  padding: 8px;
  text-align: left;
  border-bottom: 1px solid #ddd;
}
</style>""" +\
"""<center><h4 style="margin-bottom:5"; width="200px">{}</h4>""".format(state_name) + """</center>
   <div>
     <ol>""" + bills_html_links + """</ol>
   </div>
</html>
"""
  return html


def get_popups():

  # Load links for reference
  links = load_links()

  # Holds info on each states centroid
  with open('us-state-centroids.json', 'r') as f:
    centroids = json.load(f)

  popups = []

  # Make list of popups
  for feat in centroids['features']:

    # Get state and coords
    coords = feat['geometry']['coordinates']
    state = feat['properties']['name']

    # Get HTML for pop up
    try:
        html = to_popup_full_html(state, links)
    except KeyError:
        continue

    # Generate popup and marker
    popup = folium.Popup(folium.Html(html, script=True), max_width=500)
    
    # Make markers
    marker_icon = folium.Icon(color='red', icon='university', prefix='fa')
    marker = folium.Marker([coords[1], coords[0]], popup=popup, icon=marker_icon)
    popups.append(marker)

  return popups

def make_markers_map(summed_all):

    # Init map
    m = folium.Map(location=[40, -102],
                    tiles=None,
                    zoom_start=4)

    # Add base layer
    folium.TileLayer(tiles='OpenStreetMap', overlay=True, control=False).add_to(m)

    # Set colorscale
    colorscale = branca.colormap.linear.YlOrRd_09.scale(0, summed_all['Num Bills'].max())

    # Plot just base number of bills
    style_func = get_style_function(summed_all, 'Num Bills', colorscale)

    # Get reference geo to plot
    geo = load_geo(summed_all)

    # Make plot
    folium.TopoJson(geo, 'objects.states',
                    style_function=style_func,
                    name='Num Bills',
                    tooltip=None,
                    smooth_factor=.5,
                    show=False,
                    overlay=False).add_to(m)

    # Add just one color bar
    colorscale.caption = 'Number of Bills'
    colorscale.add_to(m)

    # Add pop ups
    popups = get_popups()
    for p in popups:
      p.add_to(m)

    # Add title
    title_html = '<h3 align="center" style="font-size:16px"><b>Censorship Bill Explorer</b></h3>'
    m.get_root().html.add_child(folium.Element(title_html))

    # Save
    os.makedirs('docs/maps', exist_ok=True)
    m.save('docs/maps/bill_explorer.html')

def make_hover_map(ref_df, title):
  
  # Init map
  m = folium.Map(location=[40, -102],
                 tiles=None,
                 zoom_start=4)

  # Add base layer
  folium.TileLayer(tiles='CartoDB positron', overlay=True, control=False).add_to(m)

  # Set colorscale
  colorscale = branca.colormap.linear.YlOrRd_09.scale(0, ref_df['Num Bills'].max())

  # Get reference geo
  geo = load_geo(ref_df)

  def plot_column(col):

    # Get style func, specific to this column / layer
    style_func = get_style_function(ref_df, col, colorscale)

    # Make hover tool tip
    tooltip = folium.GeoJsonTooltip(fields=['info'], sticky=True)

    # This method adds one layer
    # To add more than one layer, we will want to
    # pass different versions of the style_function
    # as well as potentially the tooltip
    folium.TopoJson(geo, 'objects.states',
                    style_function=style_func,
                    name=col,
                    tooltip=tooltip,
                    smooth_factor=.5,
                    show=False,
                    overlay=False).add_to(m)


  # Plot each column in reference
  for col in list(ref_df):
    plot_column(col)

  # Add colorbar
  colorscale.caption = 'Number of Bills'
  colorscale.add_to(m)

  # Add layer control
  folium.LayerControl(collapsed=False).add_to(m)

  # Add title
  title_html = f'<h3 align="center" style="font-size:16px"><b>{title}</b></h3>'
  m.get_root().html.add_child(folium.Element(title_html))

  # Save under title
  os.makedirs('docs/maps', exist_ok=True)
  m.save(f"docs/maps/{title.lower().replace(' ', '_').replace('(', '').replace(')', '')}.html")

  return m

def save_in_iframe(loc, save_loc=None,
                   height='45em', width='100%',
                   style_str=None):
    
    # Over-write existing by default
    if save_loc is None:
        save_loc = loc
        
    # If passed style str, use default
    if style_str is None:
        style_str = f"border:none; height:{height}; width:{width};"
    
    # Load as html
    with open(loc, 'rb') as f:
        html = f.read()
    
    # Encode in a kind of annoying way, but makes
    # it play nice with IFrame's
    encoded = base64.b64encode(html).decode('utf8')
    src = "data:text/html;charset=utf-8;base64," + encoded
    
    # Over-write the existing file, but wrapped in an iframe
    with open(save_loc, 'w') as f:
        f.write(f'<iframe src="{src}" scrolling="no" style="{style_str}"></iframe>')

def main():

    # Download
    dl()

    # Get dataframe
    df = get_df()

    # Base sums
    summed_all = sum_df_by_state(df)

    # Sum by state x group combos
    summed_by_status = sum_df_by_state(df, 'Refined Status')
    summed_by_year = sum_df_by_state(df, 'Year')

    # Make and save markers map
    make_markers_map(summed_all)

    # Breakdowns
    targets_cols = ['Num Bills', 'K12', 'Colleges', 'State Contractors / Institutions', 'Gender / Sexuality', 'Religion', 'Race / Ethnicity']
    status_cols = ['Num Bills', 'Num Pending', 'Num Died', 'Num Signed']

    # Make rest of maps
    make_hover_map(summed_all[targets_cols], title='Bill Targets')

    make_hover_map(summed_by_year.loc[2021][targets_cols], title='Bill Targets 2021')
    make_hover_map(summed_by_year.loc[2022][targets_cols], title='Bill Targets 2022')

    make_hover_map(summed_by_status.loc['Died'][targets_cols], title='(Dead) Bill Targets')
    make_hover_map(summed_by_status.loc['Pending'][targets_cols], title='(Pending) Bill Targets')
    make_hover_map(summed_by_status.loc['Signed'][targets_cols], title='(Signed) Bill Targets')

    make_hover_map(summed_all[status_cols], title='Bills by Status')
    make_hover_map(summed_by_year.loc[2021][status_cols], title='Bills by Status 2021')
    make_hover_map(summed_by_year.loc[2022][status_cols], title='Bills by Status 2022')

    # Add iframe version for each map saved in includes
    os.makedirs('docs/_includes', exist_ok=True)

    for file in os.listdir('docs/maps'):
      loc = os.path.join('docs/maps', file)
      save_loc = os.path.join('docs/_includes', file)
      save_in_iframe(loc, save_loc)


if __name__ == '__main__':
    main()

